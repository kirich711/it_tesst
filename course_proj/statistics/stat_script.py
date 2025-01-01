import time
import requests
import warnings
import pandas as pd
import requests
from pandas.errors import SettingWithCopyWarning
from lxml.etree import XMLSyntaxError
import openpyxl as op
from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Border, Side
from collections import Counter
import threading
import os
from concurrent.futures import ProcessPoolExecutor
from concurrent.futures import wait
import time
from multiprocessing import Manager

warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.simplefilter(action='ignore', category=SettingWithCopyWarning)


def change_date_format(_string):
    temp = str(_string).split("T")
    temp_ymd = temp[0].split("-")
    return temp_ymd


def get_df(filename):
    return pd.read_csv(filename, low_memory=False)


def prepare_df(m_df):
    def get_value_for_curr(date, curr):
        subprefix = "[PREPARING][CHANGING CURRENCY]"
        currs_dict = {}

        def get_response(m_URL):
            response = requests.get(m_URL)
            try:
                df_ = pd.read_xml(response.text)
            except XMLSyntaxError:
                print(f"{subprefix} Sleeping for 70 sec because of DDOS protection")
                print(f"{subprefix} Process will continue after sleeping")
                time.sleep(70)
                response = requests.get(m_URL)
                df_ = pd.read_xml(response.text)
            m_dct = df_.groupby('CharCode')[['VunitRate', 'Value', 'Nominal']].first().to_dict()
            return m_dct

        def get_dct(m_date):
            m_URL = f"http://www.cbr.ru/scripts/XML_daily.asp?date_req={m_date}"
            m_dct = get_response(m_URL)
            return m_dct

        def get_result_curr_value(m_date, m_curr):
            return currs_dict[m_date]["VunitRate"][m_curr]

        year = date[0]
        month = date[1]
        day = date[2]
        date_s = f"01/{month}/{year}"
        result = ""
        if not (date_s in currs_dict):
            print(f"{subprefix} Date:", date_s, "not found => Adding to dictionary")
            dct = get_dct(date_s)
            currs_dict[date_s] = dct

        try:
            result = get_result_curr_value(date_s, curr)
        except KeyError:
            if curr == 'GEL':
                return None
            if curr == "BYR":
                curr = "BYN"
                result = get_result_curr_value(date_s, curr)

        return result

    def change_curr(m_df):
        curr = m_df['salary_currency']
        m_avg = m_df["salary_average"]
        if curr != "RUR":
            date = change_date_format(m_df['published_at'])
            str_value = get_value_for_curr(date, curr)
            if str_value is None:
                return None
            value = float(str(str_value).replace(",", '.'))
            m_avg = m_avg * value
        return m_avg

    def get_year_from_date(date):
        return change_date_format(date)[0]

    prefix = "[PREPARING]"
    print(f"{prefix} START")
    print("[PREPARING] Splitting date and calculating avg salary ...")
    m_df = (m_df
            .assign(published_at_year=lambda df_: df_['published_at'].apply(get_year_from_date).astype(int))
            .assign(salary_average=lambda df_: df_[['salary_from', 'salary_to']].mean(axis=1))
            )
    m_df_without_nan_curr = m_df.copy()
    print(f"{prefix} Dropping nan currencies...")
    m_df_without_nan_curr.dropna(subset=['salary_currency'], inplace=True)
    # m_df_without_nan_curr.to_csv('v1.csv', index=False)
    print(f"{prefix} Changing avg salary currency... (May be long)")
    m_df_without_nan_curr["salary_average"] = m_df_without_nan_curr.apply(lambda row: change_curr(row), axis=1)
    # m_df_without_nan_curr.to_csv('v2.csv', index=False)
    print(f"{prefix} Dropping nan avg salary...")
    m_df_without_nan_curr.dropna(subset=['salary_average'], inplace=True)
    # m_df_without_nan_curr.to_csv('v3.csv', index=False)
    print(f"{prefix} END")
    return m_df, m_df_without_nan_curr


def get_avg_salary_by_year(m_df, years):
    return (m_df
            .groupby('published_at_year')['salary_average']
            .mean()
            .apply(lambda x: int(round(x)))
            .reindex(years, fill_value=0)
            )


def get_vac_num_by_year(m_df, years):
    return (m_df
            .groupby('published_at_year')
            .size()
            .reindex(years, fill_value=0)
            )


def get_avg_salary_by_city(m_df):
    return (m_df
            .groupby('area_name')["salary_average"]
            .mean()
            .sort_values(ascending=False)
            .apply(lambda x: int(round(x)))
            )


def get_vac_num_by_city(m_df):
    return (m_df['area_name']
            .value_counts(normalize=True)
            .apply(lambda x: round(x, 4) * 100)
            )


def get_sorted_dict(m_dict):
    return dict(sorted(m_dict.items(), key=lambda x: (-x[1], x[0]))[0:10])


def split_key_skills(string):
    return string.replace("\r\n", "%").replace("\n", "%").replace("\r", "%").split("%")


def check_key_skills_and_name(m_row, skills_for_vac):
    name = str(m_row["name"])
    flag = False
    if "javascript" in name.lower():
        return False
    if not (pd.isna(m_row["key_skills"])):
        key_skills = split_key_skills(str(m_row["key_skills"]))
        for skill in skills_for_vac:
            if skill in key_skills:
                flag = True
    for skill in skills_for_vac:
        if skill in name.lower():
            flag = True
    return flag


def count_skills(m_df, num_workers=6):
    def get_counted_skills(m_row):
        key_skills = split_key_skills(str(m_row["key_skills"]))
        count = Counter(key_skills)
        return dict(count)

    def worker_job(group, year, m_result_dict):
        pid = os.getpid()
        thread = threading.current_thread()
        worker_prefix = f"[COUNTING][GROUPING] WORKER ({thread.name} of process with id:{pid}) "
        print(f"{worker_prefix} STARTS with {year}")
        dct = dict(Counter(group["skills_count"].sum()).most_common(20))
        m_result_dict[year] = dct
        print(f"{worker_prefix} ENDS with {year}")

    prefix = "[COUNTING]"
    print(f"{prefix} START")
    print(f"{prefix} Dropping nan skills...")
    m_df = m_df[["published_at_year", "key_skills"]]
    m_df.dropna(subset=['key_skills'], inplace=True)
    print(f"{prefix} Adding counted skills...")
    # m_df.to_csv('skills_1.csv', index=False)
    m_df["skills_counted"] = m_df.apply(lambda row: get_counted_skills(row), axis=1)
    m_df["skills_count"] = m_df["skills_counted"].apply(lambda x: Counter(x))
    # m_df.to_csv('skills_2.csv', index=False)
    print(f"{prefix} Grouping counted skills by years... (May be long)")
    m_df.index = m_df.index.astype(int)
    m_df['published_at_year'] = m_df['published_at_year'].astype(int)
    groups = m_df.groupby('published_at_year', sort=False)

    poolexec = ProcessPoolExecutor(max_workers=num_workers)

    with Manager() as manager:
        shared_dict = manager.dict()
        with poolexec as exec:
            fs = [exec.submit(worker_job, group, yr, shared_dict) for yr, group in groups]
            wait(fs)
            result_dict = dict(shared_dict)
            min_year = min(shared_dict.keys())
            max_year = max(shared_dict.keys())
            return result_dict, min_year, max_year


def create_report(years,
                  year_avg_vac_dict, year_avg_vac_dict_only_vac, sorted_salary_by_city, sorted_count_by_city,
                  sorted_count_by_city_only_vac, sorted_salary_by_city_only_vac,
                  dict_counted_skills, dict_counted_skills_vac, min_year, max_year):
    side = Side(style="thin")
    border = Border(left=side, right=side, top=side, bottom=side)
    font = Font(bold=True)
    alignment_right = Alignment(horizontal='right')
    alignment_left = Alignment(horizontal='left')

    def set_title(m_sheet, m_dict):
        for (m_key, m_value) in m_dict.items():
            m_sheet[f"{m_key}1"] = m_value
            m_sheet[f"{m_key}1"].font = font

    def print_sheet_area(m_sheet, m_dict_items, area_letter, value_letter):
        i = 2
        for (m_area, m_value) in m_dict_items:
            m_sheet[f"{area_letter}{i}"] = m_area
            m_sheet[f"{value_letter}{i}"] = m_value
            m_sheet[f"{area_letter}{i}"].alignment = alignment_left
            m_sheet[f"{value_letter}{i}"].alignment = alignment_right
            m_sheet[f"{area_letter}{i}"].border = border
            m_sheet[f"{value_letter}{i}"].border = border
            i += 1

    def print_sheet_skill(m_sheet, m_dict, sheet_dict):
        del sheet_dict["A"]
        print(sheet_dict.keys())
        for (m_letter, m_year) in sheet_dict.items():
            i = 2
            dct = m_dict[int(m_year)]
            for (skill, count) in dct.items():
                m_sheet[f"{m_letter}{i}"] = str(skill)
                m_sheet[f"{m_letter}{i}"].offset(row=0, column=1).value = count
                m_sheet[f"{m_letter}{i}"].alignment = alignment_left
                m_sheet[f"{m_letter}{i}"].offset(row=0, column=1).alignment = alignment_right
                m_sheet[f"{m_letter}{i}"].border = border
                m_sheet[f"{m_letter}{i}"].offset(row=0, column=1).border = border
                i += 1

    def print_sheet_year(m_sheet, m_dict, year_letter, avg_letter, num_letter):
        i = 2
        for m_year in years:
            m_sheet[f"{year_letter}{i}"] = m_year
            m_sheet[f"{avg_letter}{i}"] = m_dict["sal_avg"][m_year]
            m_sheet[f"{num_letter}{i}"] = m_dict["vac_num"][m_year]
            m_sheet[f"{year_letter}{i}"].alignment = alignment_right
            m_sheet[f"{avg_letter}{i}"].alignment = alignment_right
            m_sheet[f"{num_letter}{i}"].alignment = alignment_right
            m_sheet[f"{year_letter}{i}"].border = border
            m_sheet[f"{avg_letter}{i}"].border = border
            m_sheet[f"{num_letter}{i}"].border = border
            i += 1

    excel_doc = op.Workbook()

    excel_doc.remove(excel_doc["Sheet"])
    excel_doc.create_sheet(title='Статистика по годам', index=0)
    excel_doc.create_sheet(title='Статистика по годам BBACKEND', index=1)
    excel_doc.create_sheet(title='Статистика по городам ', index=2)
    excel_doc.create_sheet(title='Статистика по городам BACKEND', index=3)
    excel_doc.create_sheet(title='Стат. ключ. скилл по год', index=4)
    excel_doc.create_sheet(title='Стат. ключ. скилл по год BACKEND', index=5)

    sheetnames = excel_doc.sheetnames
    sheet_year = excel_doc[sheetnames[0]]
    sheet_year_vac = excel_doc[sheetnames[1]]
    sheet_area = excel_doc[sheetnames[2]]
    sheet_area_vac = excel_doc[sheetnames[3]]
    sheet_skill = excel_doc[sheetnames[4]]
    sheet_skill_vac = excel_doc[sheetnames[5]]

    sheet_year_dict = {
        "A": "Год",
        "B": "Средняя зарплата",
        "C": "Количество вакансий"
    }

    set_title(sheet_year, sheet_year_dict)
    print_sheet_year(sheet_year, year_avg_vac_dict, "A", "B", "C")

    set_title(sheet_year_vac, sheet_year_dict)
    print_sheet_year(sheet_year_vac, year_avg_vac_dict_only_vac, "A", "B", "C")

    sheet_area_dict = {
        "A": "Город",
        "B": "Средняя зарплата",
        "D": "Город",
        "E": "Процент вакансий",
    }

    set_title(sheet_area, sheet_area_dict)
    print_sheet_area(sheet_area, sorted_salary_by_city.items(), "A", "B")
    print_sheet_area(sheet_area, sorted_count_by_city.items(), "D", "E")

    set_title(sheet_area_vac, sheet_area_dict)
    print_sheet_area(sheet_area_vac, sorted_salary_by_city_only_vac.items(), "A", "B")
    print_sheet_area(sheet_area_vac, sorted_count_by_city_only_vac.items(), "D", "E")

    letters = ["B", "E", "H", "K", "N", "Q", "T", "W", "Z"]
    skill_years = range(min_year, max_year + 1)
    sheet_skill_dict = {
        "A": "Год"
    }
    letter_i = 0
    for year in skill_years:
        sheet_skill_dict[letters[letter_i]] = str(year)
        letter_i += 1

    sheet_skill_vac_dict = {
        "A": "Год"
    }

    letter_i = 0
    for year in skill_years:
        sheet_skill_vac_dict[letters[letter_i]] = str(year)
        letter_i += 1

    set_title(sheet_skill, sheet_skill_dict)
    print_sheet_skill(sheet_skill, dict_counted_skills, sheet_skill_dict)

    set_title(sheet_skill_vac, sheet_skill_vac_dict)
    print_sheet_skill(sheet_skill_vac, dict_counted_skills_vac, sheet_skill_vac_dict)

    excel_doc.save('report.xlsx')


def build_year_avg_vac(avg_df, vac_count_df, years):
    year_avg_vac_dict = pd.DataFrame(
        {
            'sal_avg': get_avg_salary_by_year(avg_df, years),
            'vac_num': get_vac_num_by_year(vac_count_df, years)
        }).to_dict()
    return year_avg_vac_dict


def drop_less_than_percent(m_df, percent):
    m_df["percent"] = (m_df
                       .groupby('area_name')['area_name']
                       .transform(lambda x: len(x) / len(m_df))
                       )
    vdf = m_df[m_df["percent"] >= percent]
    return vdf


def input_int_field(field_name, text, err_text, err_text2):
    prefix = f'[INPUT FILED "{field_name}"]'
    flag = True
    while flag:
        field = input(f'{prefix} {text}')
        if field.isdigit():
            if int(field) > 0:
                flag = False
            else:
                print(f'{prefix} {err_text2}')
        else:
            print(f'{prefix} {err_text}')
    return int(field)


def main():
    print("START")
    year_from = 2003
    year_to = 2024
    years = range(year_from, year_to)
    percent = 0.01

    num_workers = input_int_field("num_workers",
                                  "Enter the maximum number of workers from multiprocessing (recomended 6) (depends on RAM and  processor): ",
                                  "Number of workers should be numeric!", "Number of workers should be more than 0!")

    skills_for_vac = ['backend', 'бэкэнд', 'бэкенд', 'бекенд', 'бекэнд', 'back end', 'бэк энд', 'бэк енд', 'django', 'flask', 'laravel', 'yii', 'symfony']
    csv = 'vacancies_2024.csv'

    print(f"Getting df from '{csv}'...")
    vacancies = get_df(csv)

    print("Preparing df and df_non_nan_curr...")
    vacancies, vacancies_without_nan_curr = prepare_df(vacancies)
    vacancies_without_nan_curr.to_csv('vacancies_without_nan_curr.csv', index=False)

    print("Building year_avg_vac_dict...")
    year_avg_vac_dict = build_year_avg_vac(vacancies_without_nan_curr, vacancies, years)

    print("Dropping areas less than percent = ", percent)
    vacancies_non_percent = drop_less_than_percent(vacancies, percent)
    vacancies_without_nan_curr_non_percent = drop_less_than_percent(vacancies_without_nan_curr, percent)

    print("Building count_by_city...")
    count_by_city = get_vac_num_by_city(vacancies_non_percent)
    # count_by_city.to_csv('count_by_city.csv', index=False)

    print("Building salary_by_city...")
    salary_by_city = get_avg_salary_by_city(vacancies_without_nan_curr_non_percent)
    # salary_by_city.to_csv('salary_by_city.csv', index=False)

    print("Sorting areas dfs...")
    sorted_count_by_city = get_sorted_dict(count_by_city.to_dict())
    sorted_salary_by_city = get_sorted_dict(salary_by_city.to_dict())

    print("Preparing df only vac...")
    vacancies_only_vac = vacancies[vacancies.apply(lambda row: check_key_skills_and_name(row, skills_for_vac), axis=1)]
    # #vacancies_only_vac.to_csv('v4onlyvac.csv', index=False)

    print("Counting skills...")
    dict_counted_skills, min_year, max_year = count_skills(vacancies,num_workers)
    dict_counted_skills_vac, min_year_vac, max_year_vac = count_skills(vacancies_only_vac,num_workers)
    # skill_count_by_year_only_vac.to_csv('skills_3.csv')

    print("Preparing df_non_nan_curr only vac...")
    vacancies_without_nan_curr_only_vac = vacancies_without_nan_curr[
        vacancies_without_nan_curr.apply(lambda row: check_key_skills_and_name(row, skills_for_vac), axis=1)]
    # vacancies_without_nan_curr_only_vac.to_csv('v5onlyvac.csv', index=False)

    print("Building year_avg_vac_dict only vac...")
    year_avg_vac_dict_only_vac = build_year_avg_vac(vacancies_without_nan_curr_only_vac, vacancies_only_vac, years)

    print("Dropping areas less than percent = ", percent)
    vacancies_only_vac_non_percent = drop_less_than_percent(vacancies_only_vac, percent)
    vacancies_without_nan_curr_only_vac_non_percent = drop_less_than_percent(vacancies_without_nan_curr_only_vac,
                                                                             percent)
    print("Building count_by_city only vac...")
    count_by_city_only_vac = get_vac_num_by_city(vacancies_only_vac_non_percent)
    # count_by_city_only_vac.to_csv('count_by_city only vac.csv', index=False)

    print("Building salary_by_city only vac...")
    salary_by_city_only_vac = get_avg_salary_by_city(vacancies_without_nan_curr_only_vac_non_percent)
    # salary_by_city_only_vac.to_csv('salary_by_city only vac.csv', index=False)

    print("Sorting areas dfs only vac...")
    sorted_count_by_city_only_vac = get_sorted_dict(count_by_city_only_vac.to_dict())
    sorted_salary_by_city_only_vac = get_sorted_dict(salary_by_city_only_vac.to_dict())

    print("Creating excel report...")
    create_report(years,
                  year_avg_vac_dict, year_avg_vac_dict_only_vac, sorted_salary_by_city, sorted_count_by_city,
                  sorted_count_by_city_only_vac, sorted_salary_by_city_only_vac,
                  dict_counted_skills, dict_counted_skills_vac, min_year, max_year)
    print("END")


if __name__ == '__main__':
    main()
